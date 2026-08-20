# -*- coding: utf-8 -*-
"""Разбор прайс-таблицы (xlsx) в структуру позиций калькулятора.

Ожидаемый формат близок к эталонному справочнику:
  ID | Раздел | Подраздел | Наименование позиции | Тип позиции | Единица измерения |
  Минимальная цена | Базовая цена | Максимальная цена | Статус | Основание цены | Источник | Комментарий

Разбор гибкий: колонки определяются по заголовкам, цены — по колонкам с числами.
"""
import io
import os

import openpyxl


def _norm(v):
    if v is None:
        return ''
    s = str(v).replace('\u00a0', ' ').replace('\u202f', ' ').strip()
    return s.lower()


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = (str(v)
         .replace('\u00a0', '')
         .replace('\u202f', '')
         .replace(' ', '')
         .replace(',', '.')
         .replace('₽', '')
         .replace('руб', '')
         .replace('руб.', ''))
    try:
        return round(float(s), 2)
    except (TypeError, ValueError):
        return 0.0


_SKIP_NAMES = ('итого', 'итог', 'всего', 'сумма')


def _is_skip(name):
    return _norm(name) in _SKIP_NAMES or _norm(name).startswith('итого')


def _find_header_row(rows, max_scan=25):
    for i, row in enumerate(rows[:max_scan]):
        cells = [_norm(c) for c in row]
        for c in cells:
            if 'наименован' in c or 'позиц' in c or 'раздел' in c:
                return i
    return 0


def _detect_cols(header_cells):
    cols = {}
    for i, h in enumerate(header_cells):
        hl = h.lower().strip()
        if 'наименован' in hl or (hl.startswith('наименова') and 'позиц' in hl):
            cols['name'] = i
        elif hl in ('тип', 'тип позиции', 'тип позиц', 'тип позиции,', 'тип затрат',
                    'тип работы', 'тип строки', 'категория цены') or (hl.startswith('тип ')):
            cols['type'] = i
        elif 'подраздел' in hl:
            cols['subsection'] = i
        elif 'раздел' in hl and 'под' not in hl:
            cols['section'] = i
        elif 'единиц' in hl:
            cols['unit'] = i
        elif 'минимальн' in hl:
            cols['min'] = i
        elif 'базов' in hl:
            cols['base'] = i
        elif 'максимальн' in hl:
            cols['max'] = i
        elif 'статус' in hl or 'обязательн' in hl or 'применяемость' in hl:
            cols['status'] = i
        elif 'основани' in hl:
            cols['basis'] = i
        elif 'источник' in hl or 'ссылк' in hl or 'url' in hl or 'сайт' in hl:
            cols['source'] = i
        elif 'комментар' in hl or 'примечан' in hl:
            cols['comment'] = i
    return cols


def _price_cols_from_headers(header_cells, cols):
    """Дозаполняем ценовые колонки по заголовкам, содержащим «цен»."""
    price_idx = [i for i, h in enumerate(header_cells) if 'цен' in h.lower()]
    keys = [k for k in ('min', 'base', 'max') if k not in cols]
    for i, key in zip(price_idx, keys):
        cols[key] = i


def _infer_price_cols(rows, header_row, cols):
    """Фолбэк: ценовые колонки — числовые, правее колонки единицы/названия."""
    if 'min' in cols and 'base' in cols and 'max' in cols:
        return
    anchor = cols.get('unit', cols.get('name', cols.get('section', 0)))
    # пробегаемся по строкам, собираем индексы числовых колонок после anchor
    numeric = set()
    for row in rows[header_row + 1:]:
        for i in range(anchor + 1, min(anchor + 8, len(row))):
            v = row[i]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric.add(i)
    ordered = sorted(numeric)
    keys = [k for k in ('min', 'base', 'max') if k not in cols]
    for key, i in zip(keys, ordered[:3]):
        cols[key] = i


def _parse_sheet(ws, filename):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    header_row = _find_header_row(rows)
    header_cells = rows[header_row]
    cols = _detect_cols(header_cells)
    if 'name' not in cols:
        return None
    _price_cols_from_headers(header_cells, cols)
    _infer_price_cols(rows, header_row, cols)

    items = []
    for r in rows[header_row + 1:]:
        if not any(c is not None for c in r):
            continue

        def g(key, default=''):
            idx = cols.get(key)
            if idx is None or idx >= len(r):
                return default
            return r[idx]

        name = g('name')
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        if _is_skip(name):
            continue

        prices = {k: _num(g(k, 0)) for k in ('min', 'base', 'max')}
        if prices['min'] <= 0 and prices['base'] <= 0 and prices['max'] <= 0:
            continue
        if prices['base'] <= 0:
            prices['base'] = prices['max'] or prices['min']

        items.append({
            'section': str(g('section') or '').strip(),
            'subsection': str(g('subsection') or '').strip(),
            'name': name,
            'type': str(g('type') or '').strip(),
            'unit': str(g('unit') or '').strip(),
            'min': prices['min'],
            'base': prices['base'],
            'max': prices['max'],
            'status': str(g('status') or '').strip(),
            'basis': str(g('basis') or '').strip(),
            'source': str(g('source') or '').strip(),
            'comment': str(g('comment') or '').strip(),
        })
    return items


def _collect_controls(wb):
    controls = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [_norm(h) for h in rows[0]]
        if not any('сценар' in h for h in hdr):
            continue
        idx = {k: None for k in ('scenario', 'area', 'min', 'base', 'max', 'comment', 'source')}
        for i, h in enumerate(hdr):
            if 'сценар' in h:
                idx['scenario'] = i
            elif 'площад' in h:
                idx['area'] = i
            elif 'минимальн' in h:
                idx['min'] = i
            elif 'базов' in h:
                idx['base'] = i
            elif 'максимальн' in h:
                idx['max'] = i
            elif 'комплектац' in h or 'пояснен' in h:
                idx['comment'] = i
            elif 'источник' in h:
                idx['source'] = i
        for r in rows[1:]:
            if not any(c is not None for c in r):
                continue
            scenario = r[idx['scenario']] if idx['scenario'] is not None else None
            if not scenario:
                continue
            def g(k):
                j = idx[k]
                return r[j] if j is not None and j < len(r) else None
            controls.append({
                'scenario': str(g('scenario') or '').strip(),
                'area': _num(g('area')),
                'min': _num(g('min')),
                'base': _num(g('base')),
                'max': _num(g('max')),
                'comment': str(g('comment') or '').strip(),
                'source': str(g('source') or '').strip(),
            })
    return controls


def _finalize(items, controls, source_name):
    sections = []
    for it in items:
        if it['section'] and it['section'] not in sections:
            sections.append(it['section'])
    for i, it in enumerate(items):
        it['id'] = i
    return {
        'meta': {
            'source_file': os.path.basename(source_name),
            'items_total': len(items),
            'sections_total': len(sections),
            'note': 'Данные загружены из прайс-таблицы.',
        },
        'items': items,
        'sections': sections,
        'controls': controls,
    }


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return _parse_wb(wb, os.path.basename(path))


def parse_bytes(raw, filename='upload.xlsx'):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    return _parse_wb(wb, filename)


def _parse_wb(wb, source_name):
    parsed = None
    controls = _collect_controls(wb)
    # ищем лист с позициями: тот, где нашлись наименования
    for ws in wb.worksheets:
        items = _parse_sheet(ws, source_name)
        if items and (parsed is None or len(items) > len(parsed['items'])):
            parsed = {'items': items, 'controls': controls}
    if parsed is None:
        parsed = {'items': [], 'controls': controls}
    return _finalize(parsed['items'], parsed['controls'], source_name)
