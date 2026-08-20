# -*- coding: utf-8 -*-
"""Экспорт сметы в Excel (xlsx)."""
import datetime
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill('solid', fgColor='16233B')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='F0E6D2')
TOTAL_FILL = PatternFill('solid', fgColor='F5A524')
TOTAL_FONT = Font(bold=True, size=12, color='1A1A1A')
THIN = Side(style='thin', color='C9D1E0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')
RIGHT = Alignment(horizontal='right', vertical='top')


def _num_fmt(v):
    if v is None:
        return 0.0
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def _write_row(ws, row, values, *, bold=False, fill=None, font_color=None, number_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        c.alignment = WRAP
        if bold:
            c.font = Font(bold=True)
        if font_color:
            c.font = Font(bold=bold, color=font_color)
        if fill:
            c.fill = fill
        if i in number_cols and isinstance(v, (int, float)):
            c.number_format = '#,##0'
            c.alignment = RIGHT


def build(payload):
    params = payload.get('params', {})
    items = payload.get('items', [])
    area = _num_fmt(params.get('area'))
    client = str(params.get('client') or '')
    obj = str(params.get('object') or '')
    comment = str(params.get('comment') or '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Смета'
    ws.sheet_view.showGridLines = False

    widths = [6, 22, 46, 12, 12, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_row = 1
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
    tcell = ws.cell(row=title_row, column=1, value='СМЕТА НА СТРОИТЕЛЬСТВО КАРКАСНОГО ДОМА')
    tcell.font = Font(bold=True, size=14, color='FFFFFF')
    tcell.fill = TOTAL_FILL
    tcell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[title_row].height = 26

    meta_rows = [('Площадь дома, м²', area), ('Дата расчёта', datetime.date.today().isoformat())]
    if obj:
        meta_rows.append(('Объект', obj))
    if client:
        meta_rows.append(('Заказчик', client))
    if comment:
        meta_rows.append(('Комментарий', comment))
    r = title_row + 1
    for k, v in meta_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    header_row = r + 1

    headers = ['№', 'Раздел', 'Наименование позиции', 'Ед. изм.', 'Кол-во',
               'Цена, ₽', 'Сумма, ₽', 'Статус']
    _write_row(ws, header_row, headers, bold=True, fill=HDR_FILL, font_color='FFFFFF')
    for col in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col).alignment = CENTER
        ws.cell(row=header_row, column=col).font = HDR_FONT

    current_section = None
    section_start = None
    subtotal = 0.0
    total = 0.0
    row = header_row + 1
    idx = 0

    def flush_section():
        nonlocal subtotal, row, section_start, current_section
        if current_section is not None and section_start is not None and section_start != row:
            _write_row(ws, row, ['', f'Итого по разделу «{current_section}»', '', '', '',
                                 '', round(subtotal, 2), ''],
                       bold=True, fill=SUB_FILL)
            ws.cell(row=row, column=7).font = Font(bold=True)
            row += 1
        subtotal = 0.0
        section_start = None

    for it in items:
        if not it.get('included'):
            continue
        section = str(it.get('section') or 'Прочее')
        if section != current_section:
            flush_section()
            current_section = section
            section_start = row
        idx += 1
        qty = it.get('qty')
        price = _num_fmt(it.get('price'))
        amount = _num_fmt(it.get('amount'))
        subtotal += amount
        total += amount
        qty_disp = qty if isinstance(qty, (int, float)) and qty >= 0 else 0
        _write_row(ws, row, [
            idx, section, it.get('name', ''), it.get('unit', ''),
            round(qty_disp, 2), round(price, 2), round(amount, 2),
            it.get('status', ''),
        ], number_cols=(5, 6, 7))
        row += 1

    flush_section()
    _write_row(ws, row, ['', 'ИТОГО, ₽', '', '', '', '', round(total, 2), ''],
               bold=True, fill=TOTAL_FILL, font_color='1A1A1A')
    ws.cell(row=row, column=7).number_format = '#,##0'
    total_row = row

    # Итоги по разделам
    ws2 = wb.create_sheet('Итоги по разделам')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 10
    _write_row(ws2, 1, ['Раздел', 'Сумма, ₽', 'Доля, %'], bold=True, fill=HDR_FILL, font_color='FFFFFF')
    sections = {}
    for it in items:
        if it.get('included'):
            sec = str(it.get('section') or 'Прочее')
            sections[sec] = sections.get(sec, 0.0) + _num_fmt(it.get('amount'))
    rr = 2
    for sec, sm in sections.items():
        share = (sm / total * 100) if total else 0.0
        _write_row(ws2, rr, [sec, round(sm, 2), round(share, 1)], number_cols=(2,))
        rr += 1
    _write_row(ws2, rr, ['ИТОГО', round(total, 2), 100.0], bold=True, fill=TOTAL_FILL, font_color='1A1A1A')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
