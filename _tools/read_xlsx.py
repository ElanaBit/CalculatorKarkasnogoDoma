# -*- coding: utf-8 -*-
import glob, sys, io
import openpyxl
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

files = glob.glob('*.xlsx')
wb = openpyxl.load_workbook(files[0], data_only=True)
for ws in wb.worksheets:
    print('SHEET:', ws.title, 'dims:', ws.dimensions, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
    print('hidden cols:', [k for k, v in ws.column_dimensions.items() if v.hidden])
    print('hidden rows:', [k for k, v in ws.row_dimensions.items() if v.hidden])
    print('merged:', ws.merged_cells.ranges)
