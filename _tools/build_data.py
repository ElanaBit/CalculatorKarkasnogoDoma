# -*- coding: utf-8 -*-
"""Генерирует embedded_data.json из эталонного xlsx (справочник затрат)."""
import glob, json, os, sys, io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

files = glob.glob(os.path.join(ROOT, '*.xlsx'))
assert files, 'Не найден xlsx файл'
wb = openpyxl.load_workbook(files[0], data_only=True)

price_ws = wb['Справочник затрат']
rows = list(price_ws.iter_rows(values_only=True))
header = [str(h).strip() if h is not None else '' for h in rows[0]]

def col(name):
    for i, h in enumerate(header):
        if name.lower() in h.lower():
            return i
    return None

I_SEC = col('раздел')
I_SUB = col('подраздел')
I_NAME = col('наименование')
I_TYPE = col('тип')
I_UNIT = col('единица')
I_MIN = col('минимальная')
I_BASE = col('базовая')
I_MAX = col('максимальная')
I_STATUS = col('статус')
I_BASIS = col('основание')
I_SOURCE = col('источник')
I_COMMENT = col('комментарий')

items = []
for r in rows[1:]:
    if not any(r):
        continue
    def g(idx):
        return r[idx] if idx is not None and idx < len(r) else None
    name = g(I_NAME)
    if name is None or not str(name).strip():
        continue
    def num(v):
        try:
            f = float(v)
            return round(f, 2)
        except (TypeError, ValueError):
            return 0.0
    item = {
        'section': str(g(I_SEC) or '').strip(),
        'subsection': str(g(I_SUB) or '').strip(),
        'name': str(name).strip(),
        'type': str(g(I_TYPE) or '').strip(),
        'unit': str(g(I_UNIT) or '').strip(),
        'min': num(g(I_MIN)),
        'base': num(g(I_BASE)),
        'max': num(g(I_MAX)),
        'status': str(g(I_STATUS) or '').strip(),
        'basis': str(g(I_BASIS) or '').strip(),
        'source': str(g(I_SOURCE) or '').strip(),
        'comment': str(g(I_COMMENT) or '').strip(),
    }
    items.append(item)

sections = []
for it in items:
    if it['section'] and it['section'] not in sections:
        sections.append(it['section'])

# Контрольные бюджеты
controls = []
ws2 = wb['Контрольные бюджеты']
rows2 = list(ws2.iter_rows(values_only=True))
hdr2 = [str(h).strip() if h is not None else '' for h in rows2[0]]
def col2(name):
    for i, h in enumerate(hdr2):
        if name.lower() in h.lower():
            return i
    return None
I_SC = col2('сценарий')
I_AR = col2('площадь')
I_MN = col2('минимальная')
I_BA = col2('базовая')
I_MX = col2('максимальная')
I_CM = col2('комплектация')
I_SR = col2('источник')
for r in rows2[1:]:
    if not any(r):
        continue
    def num2(v):
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return 0.0
    controls.append({
        'scenario': str(r[I_SC] or '').strip(),
        'area': num2(r[I_AR]) if I_AR is not None else 0,
        'min': num2(r[I_MN]) if I_MN is not None else 0,
        'base': num2(r[I_BA]) if I_BA is not None else 0,
        'max': num2(r[I_MX]) if I_MX is not None else 0,
        'comment': str(r[I_CM] or '').strip(),
        'source': str(r[I_SR] or '').strip(),
    })

data = {
    'meta': {
        'source_file': os.path.basename(files[0]),
        'items_total': len(items),
        'sections_total': len(sections),
        'note': 'Встроенный справочник. Можно обновить кнопкой «Загрузить данные».',
    },
    'items': items,
    'sections': sections,
    'controls': controls,
}

out = os.path.join(ROOT, 'app', 'embedded_data.json')
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=1)

print('Sections:', len(sections))
print('Items:', len(items))
print('Controls:', len(controls))
print('Saved to', out)
for s in sections:
    print(' -', s)
